import os
import win32com.client
import openpyxl
import random
from openpyxl import *
import re
import time
import ReadExcel
import scipy
import numpy
from scipy.stats import norm
import datetime
import matplotlib.pyplot as plt
import shapely.geometry


def build_floor_plan_and_bracing(SapModel, tower, all_floor_plans, all_floor_bracing, floor_num, floor_elev):
    print('Building floor plan...')
    floor_plan_num = tower.floor_plans[floor_num-1]
    floor_plan = all_floor_plans[floor_plan_num-1]
    floor_bracing_num = tower.floor_bracing_types[floor_num-1]
    floor_bracing = all_floor_bracing[floor_bracing_num-1]
    # Find scaling factors
    scaling_x = floor_plan.scaling_x * tower.x_width
    scaling_y = floor_plan.scaling_y * tower.y_width
    #Create members for floor plan
    for member in floor_plan.members:
        kip_in_F = 3
        SapModel.SetPresentUnits(kip_in_F)
        start_node = member.start_node
        end_node = member.end_node
        start_x = start_node[0] * scaling_x
        start_y = start_node[1] * scaling_y
        start_z = floor_elev
        end_x = end_node[0] * scaling_x
        end_y = end_node[1] * scaling_y
        end_z = floor_elev
        section_name = member.sec_prop
        [ret, name] = SapModel.FrameObj.AddByCoord(start_x, start_y, start_z, end_x, end_y, end_z, '', PropName=section_name)
        if ret != 0:
            print('ERROR creating floor plan member on floor ' + str(floor_num))
    # Create members for floor bracing
    # Create floor bracing
    print('Building floor bracing...')
    for member in floor_bracing.members:
        kip_in_F = 3
        SapModel.SetPresentUnits(kip_in_F)
        start_node = member.start_node
        end_node = member.end_node
        start_x = start_node[0] * scaling_x
        start_y = start_node[1] * scaling_y
        start_z = floor_elev
        end_x = end_node[0] * scaling_x
        end_y = end_node[1] * scaling_y
        end_z = floor_elev
        section_name = member.sec_prop
        [ret, name] = SapModel.FrameObj.AddByCoord(start_x, start_y, start_z, end_x, end_y, end_z, '', PropName=section_name)
        if ret != 0:
            print('ERROR creating floor bracing member on floor ' + str(floor_num))
    # Assign masses to mass nodes and create steel rod
    floor_mass = tower.floor_masses[floor_num-1]
    # If at the top floor of tower, assign masses to floor bracing. Else, assign to floor plan and create steel rod
    if floor_num != len(Tower.floor_plans):
        mass_nodes = floor_plan.mass_nodes
        mass_per_node = floor_mass / len(mass_nodes)
    else:
        mass_nodes = floor_bracing.mass_nodes
        mass_per_node = floor_mass / len(mass_nodes)
    # Create the mass node point
    for mass_node in mass_nodes:
        kip_in_F = 3
        SapModel.SetPresentUnits(kip_in_F)
        [ret, mass_name] = SapModel.PointObj.AddCartesian(mass_node[0]* scaling_x,mass_node[1]*scaling_y,floor_elev,MergeOff=False)
        if ret != 0:
            print('ERROR setting mass nodes on floor ' + str(floor_num))
        #Assign masses to the mass nodes
        #Shaking in the x direcion!
        N_m_C = 10
        SapModel.SetPresentUnits(N_m_C)
        ret = SapModel.PointObj.SetMass(mass_name, [mass_per_node,0,0,0,0,0],0,True,False)
        if ret[0] != 0:
            print('ERROR setting mass on floor ' + str(floor_num))
        # Create floor load forces
        N_m_C = 10
        SapModel.SetPresentUnits(N_m_C)
        ret = SapModel.PointObj.SetLoadForce(mass_name, 'DEAD', [0, 0, -1 * mass_per_node * 9.81, 0, 0, 0])

    #Create steel rod
    if floor_num != len(Tower.floor_plans):
        kip_in_F = 3
        SapModel.SetPresentUnits(kip_in_F)
        [ret, name1] = SapModel.FrameObj.AddByCoord(mass_nodes[0][0]*scaling_x, mass_nodes[0][1]*scaling_y, floor_elev, mass_nodes[1][0]*scaling_x, mass_nodes[1][1]*scaling_y, floor_elev, '', PropName='Steel rod')
        if ret !=0:
            print('ERROR creating steel rod on floor ' + str(floor_num))

    return SapModel


def build_face_bracing(SapModel, tower, all_floor_plans, all_face_bracing, floor_num, floor_elev):
    i = 1
    num_faces = len(Tower.side)
    print('Building face bracing...')
    while i <= len(Tower.side):
        face_bracing_num = Tower.bracing_types[(floor_num - 1) * num_faces + i - 1]
        if face_bracing_num != 0:
            face_bracing = all_face_bracing[face_bracing_num-1]

            #Find scaling factors
            floor_plan_num = tower.floor_plans[floor_num-1]
            floor_plan = all_floor_plans[floor_plan_num-1]
       
            scaling_x = floor_plan.scaling_x * tower.x_width
            scaling_y = floor_plan.scaling_y * tower.y_width
            scaling_z = tower.floor_heights[floor_num-1]
        
            for member in face_bracing.members:
                kip_in_F = 3
                SapModel.SetPresentUnits(kip_in_F)
                start_node = member.start_node
                end_node = member.end_node
            
                #Create face bracing for long side
                if i == 1 or i == 3:
                    scaling_x_or_y = scaling_x
                #Create face bracing for short side
                elif i == 2 or i == 4:
                    scaling_x_or_y = scaling_y

                start_x = start_node[0] * scaling_x_or_y
                start_y = 0
                start_z = start_node[1] * scaling_z + floor_elev
                end_x = end_node[0] * scaling_x_or_y
                end_y = 0
                end_z = end_node[1] * scaling_z + floor_elev
                section_name = member.sec_prop 
                #rotate coordinate system through side 1 - 4
                if i == 1:
                    ret = SapModel.CoordSys.SetCoordSys('CSys1', 0, 0, 0, 0, 0, 0)
                elif i == 2:
                    ret = SapModel.CoordSys.SetCoordSys('CSys1', scaling_x, 0, 0, 90, 0, 0)
                elif i == 3:
                    ret = SapModel.CoordSys.SetCoordSys('CSys1', 0, scaling_y, 0, 0, 0, 0)
                elif i == 4:
                    ret = SapModel.CoordSys.SetCoordSys('CSys1', 0, 0, 0, 90, 0, 0)

                [ret, name] = SapModel.FrameObj.AddByCoord(start_x, start_y, start_z, end_x, end_y, end_z, '', section_name, ' ', 'CSys1')
                if ret != 0:
                    print('ERROR creating floor bracing member on floor ' + str(floor_num))
            i += 1
    return SapModel


def build_space_bracing(SapModel, tower, all_floor_plans, all_space_bracing, floor_num, floor_elev):
    space_bracing_num = Tower.space_bracing_types[floor_num - 1]
    if space_bracing_num != 0:
        print('Building space bracing...')
        space_bracing = all_space_bracing[space_bracing_num-1]

        #Find scaling factors
        floor_plan_num = tower.floor_plans[floor_num-1]
        floor_plan = all_floor_plans[floor_plan_num-1]
       
        scaling_x = floor_plan.scaling_x * tower.x_width
        scaling_y = floor_plan.scaling_y * tower.y_width
        scaling_z = tower.floor_heights[floor_num-1]
        
        for member in space_bracing.members:
            kip_in_F = 3
            SapModel.SetPresentUnits(kip_in_F)
            start_node = member.start_node
            end_node = member.end_node

            start_x = start_node[0] * scaling_x
            start_y = start_node[1] * scaling_y
            start_z = floor_elev
            end_x = end_node[0] * scaling_x
            end_y = end_node[1] * scaling_y
            end_z = scaling_z + floor_elev
            section_name = member.sec_prop

            [ret, name] = SapModel.FrameObj.AddByCoord(start_x, start_y, start_z, end_x, end_y, end_z, '', PropName=section_name)
            if ret != 0:
                print('ERROR creating space bracing member on floor ' + str(floor_num))
    return SapModel


def build_columns(SapModel, tower, all_floor_plans, all_sections, floor_num, floor_height, floor_elev):
    print('Building columns...')
    floor_plan_num = tower.floor_plans[floor_num-1]
    floor_plan = all_floor_plans[floor_plan_num-1]
    num_corners = len(tower.side)
    x_values = []
    y_values = []
    for member in floor_plan.members:
        start_node = member.start_node
        x_values.append(start_node[0])
        y_values.append(start_node[1])   
    kip_in_F = 3
    SapModel.SetPresentUnits(kip_in_F)
    min_x = min(x_values) * tower.x_width
    max_x = max(x_values) * tower.x_width
    min_y = min(y_values) * tower.y_width
    max_y = max(y_values) * tower.y_width

    section_num = tower.col_props[(floor_num - 1) * num_corners]
    section_name = all_sections['Section ' + str(section_num)]['Name']
    [ret, name] = SapModel.FrameObj.AddByCoord(min_x, min_y, floor_elev, min_x, min_y, floor_elev + floor_height, '', PropName=section_name)
    if ret != 0:
        print('ERROR creating column on floor ' + str(floor_num))
    section_num = tower.col_props[(floor_num - 1) * num_corners + 1]
    section_name = all_sections['Section ' + str(section_num)]['Name']
    [ret, name] = SapModel.FrameObj.AddByCoord(min_x, max_y, floor_elev, min_x, max_y, floor_elev + floor_height, '', PropName=section_name)
    if ret != 0:
        print('ERROR creating column on floor ' + str(floor_num))
    section_num = tower.col_props[(floor_num - 1) * num_corners + 2]
    section_name = all_sections['Section ' + str(section_num)]['Name']
    [ret, name] = SapModel.FrameObj.AddByCoord(max_x, max_y, floor_elev, max_x, max_y, floor_elev + floor_height, '', PropName=section_name)
    if ret != 0:
        print('ERROR creating column on floor ' + str(floor_num))
    section_num = tower.col_props[(floor_num - 1) * num_corners + 3]
    section_name = all_sections['Section ' + str(section_num)]['Name']
    [ret, name] = SapModel.FrameObj.AddByCoord(max_x, min_y, floor_elev, max_x, min_y, floor_elev + floor_height, '', PropName=section_name)
    if ret != 0:
        print('ERROR creating column on floor ' + str(floor_num))
    return SapModel


def set_base_restraints(SapModel):
    # Set fixed ends on all ground level nodes
    node_num = 1
    [ret, number_nodes, all_node_names] = SapModel.PointObj.GetNameList()
    for node_name in all_node_names:
        [ret, x, y, z] = SapModel.PointObj.GetCoordCartesian(node_name, 0, 0, 0)
        if z == 0:
            [ret_set_restraint, ret] = SapModel.PointObj.SetRestraint(node_name, [True, True, True, True, True, True])
    return SapModel

def delete_within_panel(SapModel, Panel, members_to_keep = [], members_to_delete = []):
    max_decimal_places = 6
    members_deleted = []
    if len(members_to_delete) == 0:
        # Create vectors to define panel
        vec1_x = Panel.point1[0] - Panel.point2[0]
        vec1_y = Panel.point1[1] - Panel.point2[1]
        vec1_z = Panel.point1[2] - Panel.point2[2]
        vec2_x = Panel.point1[0] - Panel.point3[0]
        vec2_y = Panel.point1[1] - Panel.point3[1]
        vec2_z = Panel.point1[2] - Panel.point3[2]
        vec1 = [vec1_x, vec1_y, vec1_z]
        vec2 = [vec2_x, vec2_y, vec2_z]
        norm_vec = numpy.cross(numpy.array(vec1), numpy.array(vec2))

        [ret, number_members, all_member_names] = SapModel.FrameObj.GetNameList()
        # Loop through all members in model
        for member_name in all_member_names:
            # Get member coordinates
            [ret, member_pt1_name, member_pt2_name] = SapModel.FrameObj.GetPoints(member_name)
            if ret != 0:
                print('ERROR checking member ' + member_name)
            [ret, member_pt1_x, member_pt1_y, member_pt1_z] = SapModel.PointObj.GetCoordCartesian(member_pt1_name)
            if ret != 0:
                print('ERROR getting coordinate of point ' + member_pt1_name)
            [ret, member_pt2_x, member_pt2_y, member_pt2_z] = SapModel.PointObj.GetCoordCartesian(member_pt2_name)
            if ret != 0:
                print('ERROR getting coordinate of point ' + member_pt2_name)

            # Round the member coordinates
            member_pt1_x = round(member_pt1_x, max_decimal_places)
            member_pt1_y = round(member_pt1_y, max_decimal_places)
            member_pt1_z = round(member_pt1_z, max_decimal_places)
            member_pt2_x = round(member_pt2_x, max_decimal_places)
            member_pt2_y = round(member_pt2_y, max_decimal_places)
            member_pt2_z = round(member_pt2_z, max_decimal_places)

            # Check if the member is within the elevation of the panel
            panel_max_z = max(Panel.point1[2], Panel.point2[2], Panel.point3[2], Panel.point4[2])
            panel_min_z = min(Panel.point1[2], Panel.point2[2], Panel.point3[2], Panel.point4[2])
            if member_pt1_z <= panel_max_z and member_pt1_z >= panel_min_z and member_pt2_z <= panel_max_z and member_pt2_z >= panel_min_z:
                member_vec_x = member_pt2_x - member_pt1_x
                member_vec_y = member_pt2_y - member_pt1_y
                member_vec_z = member_pt2_z - member_pt1_z
                member_vec = [member_vec_x, member_vec_y, member_vec_z]

                # Check if member is in the same plane as the panel
                if numpy.dot(member_vec, norm_vec) == 0:
                    # To do this, check if the vector between a member point and a plane point is parallel to plane
                    test_vec = [member_pt1_x - Panel.point1[0], member_pt1_y - Panel.point1[1], member_pt1_z - Panel.point1[2]]
                    if numpy.dot(test_vec, norm_vec) == 0:
                        # Check if the member lies within the limits of the panel
                        # First, transform the frame of reference since Shapely only works in 2D
                        # Create unit vectors
                        ref_vec_1 = vec1
                        ref_vec_2 = numpy.cross(ref_vec_1, norm_vec)
                        # Project each point defining the panel onto each reference vector
                        panel_pt1_trans_1 = numpy.dot(Panel.point1, ref_vec_1) / numpy.linalg.norm(ref_vec_1)
                        panel_pt1_trans_2 = numpy.dot(Panel.point1, ref_vec_2) / numpy.linalg.norm(ref_vec_2)
                        panel_pt2_trans_1 = numpy.dot(Panel.point2, ref_vec_1) / numpy.linalg.norm(ref_vec_1)
                        panel_pt2_trans_2 = numpy.dot(Panel.point2, ref_vec_2) / numpy.linalg.norm(ref_vec_2)
                        panel_pt3_trans_1 = numpy.dot(Panel.point3, ref_vec_1) / numpy.linalg.norm(ref_vec_1)
                        panel_pt3_trans_2 = numpy.dot(Panel.point3, ref_vec_2) / numpy.linalg.norm(ref_vec_2)
                        panel_pt4_trans_1 = numpy.dot(Panel.point4, ref_vec_1) / numpy.linalg.norm(ref_vec_1)
                        panel_pt4_trans_2 = numpy.dot(Panel.point4, ref_vec_2) / numpy.linalg.norm(ref_vec_2)
                        # Project each point defining the member onto the reference vector
                        member_pt1 = [member_pt1_x, member_pt1_y, member_pt1_z]
                        member_pt2 = [member_pt2_x, member_pt2_y, member_pt2_z]
                        member_pt1_trans_1 = numpy.dot(member_pt1, ref_vec_1) / numpy.linalg.norm(ref_vec_1)
                        member_pt1_trans_2 = numpy.dot(member_pt1, ref_vec_2) / numpy.linalg.norm(ref_vec_2)
                        member_pt2_trans_1 = numpy.dot(member_pt2, ref_vec_1) / numpy.linalg.norm(ref_vec_1)
                        member_pt2_trans_2 = numpy.dot(member_pt2, ref_vec_2) / numpy.linalg.norm(ref_vec_2)
                        # Create shapely geometries to check if member is in the panel
                        poly_coords = [(panel_pt1_trans_1, panel_pt1_trans_2), (panel_pt2_trans_1, panel_pt2_trans_2), (panel_pt3_trans_1, panel_pt3_trans_2), (panel_pt4_trans_1, panel_pt4_trans_2)]
                        member_coords = [(member_pt1_trans_1, member_pt1_trans_2),(member_pt2_trans_1, member_pt2_trans_2)]
                        panel_shapely = shapely.geometry.Polygon(poly_coords)
                        member_shapely = shapely.geometry.LineString(member_coords)
                        # Delete member if it is inside the panel
                        if member_shapely.intersects(panel_shapely) == True and member_shapely.touches(panel_shapely) == False:
                            if member_name not in members_to_keep:
                                ret = SapModel.FrameObj.Delete(member_name, 0)
                                members_deleted.append(member_name)
                                if ret != 0:
                                    print('ERROR deleting member ' + member_name)
                                print('Deleted member ' + member_name)
    if len(members_to_delete) > 0:
        for member_name in members_to_delete:
            ret = SapModel.FrameObj.Delete(member_name, 0)
            members_deleted.append(member_name)
            if ret != 0:
                print('ERROR deleting member ' + member_name)
            print('Deleted member ' + member_name)
    return SapModel, members_deleted

def build_bracing_in_panel(SapModel, panel, bracing_scheme):
    members_built = []
    for member in bracing_scheme.members:
        start_node = member.start_node
        end_node = member.end_node
        # Scale the member start and end points to fit the panel location and dimensions
        # Get unit vectors to define the panel
        panel_vec_horiz_x = panel.point4[0] - panel.point1[0]
        panel_vec_horiz_y = panel.point4[1] - panel.point1[1]
        panel_vec_horiz_z =  panel.point4[2] - panel.point1[2]
        panel_vec_vert_x = panel.point2[0] - panel.point1[0]
        panel_vec_vert_y = panel.point2[1] - panel.point1[1]
        panel_vec_vert_z = panel.point2[2] - panel.point1[2]
        panel_vec_horiz = [panel_vec_horiz_x, panel_vec_horiz_y, panel_vec_horiz_z]
        panel_vec_vert = [panel_vec_vert_x, panel_vec_vert_y, panel_vec_vert_z]
        # Get the scaled start and end coordinates for the member
        # Translate point "horizontally" and "vertically"
        start_node_x = panel.point1[0] + start_node[0] * panel_vec_horiz[0] + start_node[1] * panel_vec_vert[0]
        start_node_y = panel.point1[1] + start_node[0] * panel_vec_horiz[1] + start_node[1] * panel_vec_vert[1]
        start_node_z = panel.point1[2] + start_node[0] * panel_vec_horiz[2] + start_node[1] * panel_vec_vert[2]
        end_node_x = panel.point1[0] + end_node[0] * panel_vec_horiz[0] + end_node[1] * panel_vec_vert[0]
        end_node_y = panel.point1[1] + end_node[0] * panel_vec_horiz[1] + end_node[1] * panel_vec_vert[1]
        end_node_z = panel.point1[2] + end_node[0] * panel_vec_horiz[2] + end_node[1] * panel_vec_vert[2]
        # Create the member
        [ret, member_name] = SapModel.FrameObj.AddByCoord(start_node_x, start_node_y, start_node_z, end_node_x, end_node_y, end_node_z, PropName=member.sec_prop)
        if ret != 0:
            print('ERROR building member in panel')
        members_built.append(member_name)
    return SapModel, members_built

def define_loading(SapModel, time_history_loc_1, time_history_loc_2, gm1_steps, gm1_intervals, gm2_steps, gm2_intervals, save_loc):
    print('Defining loading...')
    ##for GM1##
    # Define time history function
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    SapModel.Func.FuncTH.SetFromFile('GM1', time_history_loc_1, 1, 0, 1, 2, True)
    # Set the time history load case
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    SapModel.LoadCases.ModHistLinear.SetCase('GM1')
    SapModel.LoadCases.ModHistLinear.SetMotionType('GM1', 1)
    SapModel.LoadCases.ModHistLinear.SetLoads('GM1', 1, ['Accel'], ['U1'], ['GM1'], [1], [1], [0], ['Global'], [0])
    SapModel.LoadCases.ModHistLinear.SetTimeStep('GM1', gm1_steps, gm1_intervals)
    # Create load combination
    SapModel.RespCombo.Add('DEAD + GM1', 0)
    SapModel.RespCombo.SetCaseList('DEAD + GM1', 0, 'DEAD', 1)
    SapModel.RespCombo.SetCaseList('DEAD + GM1', 0, 'GM1', 1)
    ##for GM2##
    # Define time history function
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    SapModel.Func.FuncTH.SetFromFile('GM2', time_history_loc_2, 1, 0, 1, 2, True)
    # Set the time history load case
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    SapModel.LoadCases.ModHistLinear.SetCase('GM2')
    SapModel.LoadCases.ModHistLinear.SetMotionType('GM2', 1)
    SapModel.LoadCases.ModHistLinear.SetLoads('GM2', 1, ['Accel'], ['U1'], ['GM2'], [1], [1], [0], ['Global'], [0])
    SapModel.LoadCases.ModHistLinear.SetTimeStep('GM2', gm2_steps, gm2_intervals)
    # Create load combination
    SapModel.RespCombo.Add('DEAD + GM2', 0)
    SapModel.RespCombo.SetCaseList('DEAD + GM2', 0, 'DEAD', 1)
    SapModel.RespCombo.SetCaseList('DEAD + GM2', 0, 'GM2', 1)
    # Set damping ratios to 2.5%
    SapModel.LoadCases.ModHistLinear.SetDampConstant('GM1', 0.025)
    SapModel.LoadCases.ModHistLinear.SetDampConstant('GM2', 0.025)
    # Save the model
    ret = SapModel.File.Save(save_loc)
    if ret != 0:
        print('ERROR saving SAP2000 file')
    return SapModel


# Returns the max acceleration in g, max drift (displacement) in mm, and weight in pounds
def run_analysis(SapModel):
    kip_in_F = 3
    SapModel.SetPresentUnits(kip_in_F)
    #Run Analysis
    print('Computing...')
    SapModel.Analyze.RunAnalysis()
    print('Finished computing.')
    #Find nodes on the top floor
    roof_node_names = []
    print('Getting results...')
    [ret, number_nodes, all_node_names] = SapModel.PointObj.GetNameList()
    z_max = 0
    x_max = 0
    y_max = 0
    x_min = 0
    y_min = 0
    for node_name in all_node_names:
        [ret, x, y, z] = SapModel.PointObj.GetCoordCartesian(node_name, 0, 0, 0)
        x = round(x, 6)
        y = round(y, 6)
        z = round(z, 6)
        if x > x_max:
            x_max = x
        if y > y_max:
            y_max = y
        if z > z_max:
            z_max = z
        if x < x_min:
            x_min = x
        if y < y_min:
            y_min = y
    x_width = abs(x_max - x_min)
    y_width = abs(y_max - y_min)
    # Make sure we get results from a node that is at the quarter points on the top floor
    for node_name in all_node_names:
        [ret, x, y, z] = SapModel.PointObj.GetCoordCartesian(node_name, 0, 0, 0)
        x = round(x, 6)
        y = round(y, 6)
        z = round(z, 6)
        if z == z_max and (abs(x-x_min) == x_width/4 or abs(x-x_max) == x_width/4) and (abs(y-y_min) == y_width/4 or abs(y-y_max) == y_width/4):
            roof_node_names.append(node_name)
    print('Roof nodes:', roof_node_names)
    # Set units to metres
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    g = 9.81
    # Get WEIGHT
    # Get base reactions
    SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
    SapModel.Results.Setup.SetCaseSelectedForOutput('DEAD')
    # SapModel.Results.BaseReact(NumberResults, LoadCase, StepType, StepNum, Fx, Fy, Fz, Mx, My, Mz, gx, gy, gz)
    ret = SapModel.Results.BaseReact()
    if ret[0] != 0:
        print('ERROR getting weight')
    base_react = ret[7][0]
    total_weight = abs(base_react / g)
    # convert to lb
    total_weight = total_weight / 0.45359237
    results = []

    # Loop through GM1 and GM2
    for i in range(1, 3):
        SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        SapModel.Results.Setup.SetComboSelectedForOutput('DEAD + GM'+str(i), True)
        # set type to envelope
        SapModel.Results.Setup.SetOptionModalHist(1)
        # get max ACCELERATION
        # Set units to metres
        N_m_C = 10
        SapModel.SetPresentUnits(N_m_C)
        max_acc = 0
        for roof_node_name in roof_node_names:
            ret = SapModel.Results.JointAccAbs(roof_node_name, 0)
            max_and_min_acc = ret[7]
            max_pos_acc = max_and_min_acc[0]
            min_neg_acc = max_and_min_acc[1]
            if abs(max_pos_acc) >= abs(min_neg_acc):
                max_acc_node = abs(max_pos_acc)/g
            elif abs(min_neg_acc) >= abs(max_pos_acc):
                max_acc_node = abs(min_neg_acc)/g
            else:
                print('Could not find max acceleration')
            if max_acc_node > max_acc:
                max_acc = max_acc_node
        #Get joint DISPLACEMENT
        #Set units to millimetres
        N_mm_C = 9
        SapModel.SetPresentUnits(N_mm_C)
        max_disp = 0
        for roof_node_name in roof_node_names:
            ret = SapModel.Results.JointDispl(roof_node_name, 0)
            max_and_min_disp = ret[7]
            max_pos_disp = max_and_min_disp[0]
            min_neg_disp = max_and_min_disp[1]
            if abs(max_pos_disp) >= abs(min_neg_disp):
                max_disp_node = abs(max_pos_disp)
            elif abs(min_neg_disp) >= abs(max_pos_disp):
                max_disp_node = abs(min_neg_disp)
            else:
                print('Could not find max drift')
            if max_disp_node > max_disp:
                max_disp = max_disp_node
        # Get PERIOD
        ret = SapModel.Results.ModalPeriod()
        if ret[0] != 0:
            print('ERROR getting modal period')
        period = ret[5][0]
        # Get BASE SHEAR
        ret = SapModel.Results.BaseReact()
        if ret[0] != 0:
            print('ERROR getting base reaction')
        basesh = max(abs(ret[5][0]), abs(ret[5][1]))
        results.append([max_acc, max_disp, total_weight, period, basesh])
    return results


def get_costs(max_acc, max_disp, footprint, weight, floor_masses, floor_heights):
    # Subtract weights. Weight is initially in lb, convert to kg
    print('Calculating costs...')
    weight = (weight * 0.45359237 - sum(floor_masses)) / 0.45359237
    design_life = 100 #years
    construction_cost = 2000000*(weight**2)+6*(10**6)
    land_cost = 35000 * footprint
    annual_building_cost = (land_cost + construction_cost) / design_life
    equipment_cost = 15000000
    return_period_1 = 50
    return_period_2 = 300
    apeak_1 = max_acc #g's
    xpeak_1 = 100*max_disp/(sum(floor_heights) * 25.4) #% roof drift
    structural_damage_1 = scipy.stats.norm(1.5, 0.5).cdf(xpeak_1)
    equipment_damage_1 = scipy.stats.norm(1.75, 0.7).cdf(apeak_1)
    economic_loss_1 = structural_damage_1*construction_cost + equipment_damage_1*equipment_cost
    annual_economic_loss_1 = economic_loss_1/return_period_1
    structural_damage_2 = 0.5
    equipment_damage_2 = 0.5
    economic_loss_2 = structural_damage_2*construction_cost + equipment_damage_2*equipment_cost
    annual_economic_loss_2 = economic_loss_2/return_period_2
    annual_seismic_cost = annual_economic_loss_1 + annual_economic_loss_2
    return annual_building_cost, annual_seismic_cost


def write_to_excel(all_costs, all_results, save_loc):
    print('Writing all results to Excel...')
    filepath = save_loc + '/Results.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A2'].value = 'Tower #'
    ws['B1'].value = 'GM1'
    ws['B2'].value = 'Annual Building Cost + Annual Seismic Cost'
    ws['C2'].value = 'Acceleration (g)'
    ws['D2'].value = 'Displacement (mm)'
    ws['E2'].value = 'Weight (lb)'
    ws['F2'].value = 'Period (s)'
    ws['G2'].value = 'Base Shear (N)'
    ws['H2'].value = 'Annual Building Cost'
    ws['I2'].value = 'Annual Seismic Cost'
    ws['J1'].value = 'GM2'
    ws['J2'].value = 'Acceleration (g)'
    ws['K2'].value = 'Displacement (mm)'
    ws['L2'].value = 'Weight (lb)'
    ws['M2'].value = 'Period (s)'
    ws['N2'].value = 'Base Shear (N)'

    for tower_num in range(1, len(all_costs)+1):
        ws['A' + str(tower_num + 2)].value = tower_num
        # Write GM1 results
        ws['B' + str(tower_num + 2)].value = sum(all_costs[tower_num - 1]) # annual bldg cost + annual seismic cost
        ws['C' + str(tower_num + 2)].value = all_results[tower_num - 1][0][0] # acceleration
        ws['D' + str(tower_num + 2)].value = all_results[tower_num - 1][0][1] # displacement
        ws['E' + str(tower_num + 2)].value = all_results[tower_num - 1][0][2] # weight
        ws['F' + str(tower_num + 2)].value = all_results[tower_num - 1][0][3] # period
        ws['G' + str(tower_num + 2)].value = all_results[tower_num - 1][0][4] # base shear
        ws['H' + str(tower_num + 2)].value = all_costs[tower_num - 1][0] # annual bldg cost
        ws['I' + str(tower_num + 2)].value = all_costs[tower_num - 1][1] # seismic cost
        # Write GM2 results
        ws['J' + str(tower_num + 2)].value = all_results[tower_num - 1][1][0] # acceleration
        ws['K' + str(tower_num + 2)].value = all_results[tower_num - 1][1][1] # displacement
        ws['L' + str(tower_num + 2)].value = all_results[tower_num - 1][1][2] # weight
        ws['M' + str(tower_num + 2)].value = all_results[tower_num - 1][1][3] # period
        ws['N' + str(tower_num + 2)].value = all_results[tower_num - 1][1][4] # base shear
    wb.save(filepath)




#----START-----------------------------------------------------START----------------------------------------------------#



print('\n--------------------------------------------------------')
print('Autobuilder by University of Toronto Seismic Design Team')
print('--------------------------------------------------------\n')

#Read in the excel workbook
print("\nReading Excel spreadsheet...")
wb = load_workbook(r"C:\Users\kotab\OneDrive - University of Toronto\Autobuilder 2.0\Test 2019-12-30\L-shape 2019-12-29 (TEST).xlsm", data_only=True)
ExcelIndex = ReadExcel.get_excel_indices(wb, 'A', 'B', 2)

# Sections = ReadExcel.get_properties(wb,ExcelIndex,'Section')
# Materials = ReadExcel.get_properties(wb,ExcelIndex,'Material')
Bracing = ReadExcel.get_bracing(wb,ExcelIndex,'Bracing')
# FloorPlans = ReadExcel.get_floor_plans(wb,ExcelIndex)
# FloorBracing = ReadExcel.get_bracing(wb,ExcelIndex,'Floor Bracing')
# SpaceBracing = ReadExcel.get_bracing(wb,ExcelIndex,'Space Bracing')
Panels = ReadExcel.get_panels(wb, ExcelIndex)
AllTowers = ReadExcel.read_input_table(wb, ExcelIndex)
SaveLoc = ExcelIndex['Save location']
TimeHistoryLoc1 = ExcelIndex['Time history location 1']
TimeHistoryLoc2 = ExcelIndex['Time history location 2']

model_loc = r"C:\Users\kotab\OneDrive - University of Toronto\Autobuilder 2.0\L shape 2019-12-29\L shape - NoRigid (COMPUTE 1).sdb"

print('\nInitializing SAP2000 model...')
# create SAP2000 object
SapObject = win32com.client.Dispatch('SAP2000v15.SapObject')
# start SAP2000
SapObject.ApplicationStart()
# create SapModel Object
SapModel = SapObject.SapModel
# initialize model
SapModel.InitializeNewModel()
# open model
ret = SapModel.File.OpenFile(model_loc)

# For manually built models, some of the joint locations can be off by very small amounts (e.g. 1e-6 m).
# To fix this, round all coordinates down to 6 decimal places
print('Rounding coordinates...')
[ret, NumberPoints, AllPointNames] = SapModel.PointObj.GetNameList()
for PointName in AllPointNames:
    [ret, x, y, z] = SapModel.PointObj.GetCoordCartesian(PointName, 0, 0, 0)
    x = round(x,6)
    y = round(y,6)
    z = round(z,6)
    ret = SapModel.EditPoint.ChangeCoordinates_1(PointName, x, y, z, True)
    if ret != 0:
        print('ERROR rounding coordinates of point ' + PointName)


'''
# Define new materials
print("\nDefining materials...")
N_m_C = 10
SapModel.SetPresentUnits(N_m_C)
for Material, MatProps in Materials.items():
    MatName = MatProps['Name']
    MatType = MatProps['Material type']
    MatWeight = MatProps['Weight per volume']
    MatE = MatProps['Elastic modulus']
    MatPois = MatProps['Poisson\'s ratio']
    MatTherm = MatProps['Thermal coefficient']
    #Create material type
    ret = SapModel.PropMaterial.SetMaterial(MatName, MatType)
    if ret != 0:
        print('ERROR creating material type')
    #Set isotropic material proprties
    ret = SapModel.PropMaterial.SetMPIsotropic(MatName, MatE, MatPois, MatTherm)
    if ret != 0:
        print('ERROR setting material properties')
    #Set unit weight
    ret = SapModel.PropMaterial.SetWeightAndMass(MatName, 1, MatWeight)
    if ret != 0:
        print('ERROR setting material unit weight')

#Define new sections
print('Defining sections...')
kip_in_F = 3
SapModel.SetPresentUnits(kip_in_F)
for Section, SecProps in Sections.items():
    SecName = SecProps['Name']
    SecArea = SecProps['Area']
    SecTors = SecProps['Torsional constant']
    SecIn3 = SecProps['Moment of inertia about 3 axis']
    SecIn2 = SecProps['Moment of inertia about 2 axis']
    SecSh2 = SecProps['Shear area in 2 direction']
    SecSh3 = SecProps['Shear area in 3 direction']
    SecMod3 = SecProps['Section modulus about 3 axis']
    SecMod2 = SecProps['Section modulus about 2 axis']
    SecPlMod3 = SecProps['Plastic modulus about 3 axis']
    SecPlMod2 = SecProps['Plastic modulus about 2 axis']
    SecRadGy3 = SecProps['Radius of gyration about 3 axis']
    SecRadGy2 = SecProps['Radius of gyration about 2 axis']
    SecMat = SecProps['Material']
    #Create section property
    ret = SapModel.PropFrame.SetGeneral(SecName, SecMat, 0.1, 0.1, SecArea, SecSh2, SecSh3, SecTors, SecIn2, SecIn3, SecMod2, SecMod3, SecPlMod2, SecPlMod3, SecRadGy2, SecRadGy3, -1)
    if ret != 0:
        print('ERROR creating section property ' + SecName)
'''

AllCosts = []
AllResults = []
TowerNum = 1
ComputeTimes = []

# Define load cases
'''
gm1_Steps = ExcelIndex['GM1 time steps']
gm1_Intervals = ExcelIndex['GM1 time interval']
gm2_Steps = ExcelIndex['GM2 time steps']
gm2_Intervals = ExcelIndex['GM2 time interval']
SapModel = define_loading(SapModel, TimeHistoryLoc1, TimeHistoryLoc2, gm1_Steps, gm1_Intervals,
                          gm2_Steps, gm2_Intervals, SaveLoc)
'''

# Start scatter plot of FABI
plt.ion()
fig = plt.figure()
ax = plt.subplot(1,1,1)
ax.set_xlabel('Tower Number')
ax.set_ylabel('Total Cost')
xdata = []
ydata = []
ax.plot(xdata, ydata, 'ro', markersize=6)
plt.grid(True)

plt.show(block=False)

# Build all towers defined in spreadsheet
LastTower = None
MembersAddedLast = []
for Tower in AllTowers:
    MembersAdded = []
    #Unlock model
    SapModel.SetModelIsLocked(False)

    StartTime = time.time()
    print('\nBuilding tower number ' + str(TowerNum))
    print('-------------------------')

    # Delete all members within the plans and build correct bracing scheme
    kip_in_F = 3
    SapModel.SetPresentUnits(kip_in_F)

    # Get list of members to not delete
    MembersToKeep = []
    '''
    ret = SapModel.SelectObj.Group('MEMBERS TO KEEP')
    if ret == 0:
        [ret, NumberItems, ObjectTypes, ObjectNames] = SapModel.SelectObj.GetSelected()
        SapModel.SelectObj.ClearSelection()
    i = 0
    for Object in ObjectNames:
        if ObjectTypes[i] == 3:
            MembersToKeep.append(Object)
        i += 1
    '''
    if len(MembersAddedLast) != 0: #And the configuration is the same as the last tower
        print('Deleting members created in last iteration...')
        SapModel, MembersDeleted = delete_within_panel(SapModel, Panel, MembersToKeep, MembersAddedLast)

    for PanelNum in Tower.panels:
        BracingNum = Tower.panels[PanelNum]
        BracingScheme = Bracing[BracingNum - 1]
        Panel = Panels[PanelNum - 1]
        if len(MembersAddedLast) == 0: # or the configuration is different from the last tower
            print('Deleting members within panel ' + str(PanelNum) + '...')
            SapModel, MembersDeleted = delete_within_panel(SapModel, Panel, MembersToKeep)
        print('Building bracing scheme within panel ' + str(PanelNum) + "...")
        SapModel, MembersAddedPanel = build_bracing_in_panel(SapModel, Panel, BracingScheme)
        MembersAdded.extend(MembersAddedPanel)

    # Change the section properties of specified members
    print('\nChanging section properties of specified members...')
    for MemberToChange in Tower.members:
        NewSecProp = Tower.members[MemberToChange]
        print('Changed section of member ' + str(MemberToChange))
        SapModel.FrameObj.SetSection(str(MemberToChange), NewSecProp, 0)

    # Set base nodes to fixed
    SapModel = set_base_restraints(SapModel)

    # Join frame members if they are collinear, have the same section property, and don't have any mass assignments on the joint
    NumOfFrameJoins = 0
    [ret, NumberPoints, AllPointNames] = SapModel.PointObj.GetNameList()
    print('\nDeleting unnecessary joints...')
    for PointName in AllPointNames:
        [ret, NumberItems, ObjectTypes, ObjectNames, PointNumber] = SapModel.PointObj.GetConnectivity(PointName)
        if ret != 0:
            print('ERROR getting connectivity of point ' + PointName)
        if NumberItems == 2 and ObjectTypes[0] == 2 and ObjectTypes[1] == 2:
            [ret, Frame1Section, SAuto] = SapModel.FrameObj.GetSection(ObjectNames[0])
            [ret, Frame2Section, SAuto] = SapModel.FrameObj.GetSection(ObjectNames[1])
            if Frame1Section == Frame2Section:
                ret = SapModel.EditFrame.Join(ObjectNames[0], ObjectNames[1])
                if ret == 0:
                    NumOfFrameJoins += 1
    print('Joined ' + str(NumOfFrameJoins) + ' members')

    # Save the file
    SapModel.File.Save(SaveLoc + '/Tower ' + str(TowerNum))
    #Analyse tower and print results to spreadsheet
    print('\nAnalyzing tower number ' + str(TowerNum))
    print('-------------------------')
    # Run analysis and get weight, displacement, and acceleration
    # ret = SapModel.Analyze.SetSolverOption_1(0, 0, False)
    AllResults.append(run_analysis(SapModel))
    MaxAcc = AllResults[TowerNum-1][0][0]
    MaxDisp = AllResults[TowerNum-1][0][1]
    Weight = AllResults[TowerNum-1][0][2]
    #Calculate model cost
    Footprint = 144
    TotalHeight = [60] # inches
    TotalMass = [7.83] # kg
    AllCosts.append(get_costs(MaxAcc, MaxDisp, Footprint, Weight, TotalMass, TotalHeight))
    #Unlock model
    SapModel.SetModelIsLocked(False)

    '''
    # Delete everything in the model
    ret = SapModel.SelectObj.All(False)
    if ret != 0:
        print('ERROR selecting all')
    ret = SapModel.FrameObj.Delete(Name='', ItemType=2)
    if ret != 0:
        print('ERROR deleting all')
    '''
    # Determine total time taken to build current tower
    EndTime = time.time()
    TimeToComputeTower = EndTime - StartTime
    ComputeTimes.append(TimeToComputeTower)
    AverageComputeTime = sum(ComputeTimes) / len(ComputeTimes)
    ElapsedTime = sum(ComputeTimes)
    EstimatedTimeRemaining = (len(AllTowers) - TowerNum) * AverageComputeTime
    if EstimatedTimeRemaining <= 60:
        TimeUnitEstTime = 'seconds'
    elif EstimatedTimeRemaining > 60 and EstimatedTimeRemaining < 3600:
        TimeUnitEstTime = 'minutes'
        EstimatedTimeRemaining = EstimatedTimeRemaining / 60
    else:
        TimeUnitEstTime = 'hours'
        EstimatedTimeRemaining = EstimatedTimeRemaining / 3600

    if ElapsedTime <= 60:
        TimeUnitElaTime = 'seconds'
    elif ElapsedTime > 60 and ElapsedTime < 3600:
        TimeUnitElaTime = 'minutes'
        ElapsedTime = ElapsedTime / 60
    else:
        TimeUnitElaTime = 'hours'
        ElapsedTime = ElapsedTime / 3600
    #Round the times to the nearest 0.1
    AverageComputeTime = int(AverageComputeTime/1) + round(AverageComputeTime - int(AverageComputeTime/1),1)
    EstimatedTimeRemaining = int(EstimatedTimeRemaining/1) + round(EstimatedTimeRemaining - int(EstimatedTimeRemaining/1),1)
    ElapsedTime = int(ElapsedTime/1) + round(ElapsedTime - int(ElapsedTime/1), 1)

    # Add cost to scatter plot
    xdata.append(TowerNum)
    ydata.append(AllCosts[TowerNum-1][0] + AllCosts[TowerNum-1][1])
    ax.lines[0].set_data(xdata,ydata)
    ax.relim()
    ax.autoscale_view()
    plt.xticks(numpy.arange(min(xdata), max(xdata)+1, 1.0))
    plt.title('Average time per tower: ' + str(AverageComputeTime) + ' seconds\n' + 'Estimated time remaining: ' + str(EstimatedTimeRemaining) + ' ' + TimeUnitEstTime + '\nElapsed time so far: ' + str(ElapsedTime) + ' ' + TimeUnitElaTime)
    fig.canvas.flush_events()

    # Increment tower number
    TowerNum += 1
    LastTower = Tower
    MembersAddedLast = MembersAdded

print('\n\nFinished constructing all towers.')

# Write all results to excel spreadsheet
write_to_excel(AllCosts, AllResults, SaveLoc)
# Close SAP2000
print('Closing SAP2000...')
SapObject.ApplicationExit(False)
print('FINISHED.')
plt.show(block=True)

